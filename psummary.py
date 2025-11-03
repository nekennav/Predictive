import streamlit as st
import pandas as pd
from datetime import datetime
import os
import io
import shutil
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from werkzeug.utils import secure_filename

# -------------------------------------------------
# Page config + Space-theme CSS
# -------------------------------------------------
st.set_page_config(page_title="NEIL", page_icon="Gem", layout="wide")
st.markdown(
    """
    <style>
    .stApp {
        background-image: url('https://images.wallpaperscraft.com/image/single/planet_space_universe_127497_2560x1440.jpg');
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-color: rgba(10,15,30,0.95);
        color: #e0e0e0;
        animation: panningBackground 30s linear infinite;
    }
    @keyframes panningBackground {
        0% { background-position: 50% 0% }
        50% { background-position: 50% 100% }
        100% { background-position: 50% 0% }
    }
    h1 {
        color: #a6b1e1;
        font-family: Arial, sans-serif;
        text-align: center;
        text-shadow: -1px -1px 0 #fff, 1px -1px 0 #fff, -1px 1px 0 #fff, 1px 1px 0 #fff,
                     0 0 10px rgba(166,177,225,0.8);
        font-size: 2.5rem;
    }
    h2 { color: #dcd6ff; font-size: 1.5rem; }
    .stFileUploader {
        background: rgba(162,138,255,0.1);
        border: 2px solid #a28aff;
        border-radius: 10px;
        padding: 10px;
        margin-top: 20px;
    }
    .stFileUploader label { color: #e0e0e0; font-size: 1.1rem; }
    .stButton>button {
        background: #a28aff;
        color: #e0e0e0;
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: bold;
        border: 1px solid #dcd6ff;
        font-size: 1rem;
    }
    .stButton>button:hover { background: #8b74e6; border-color: #a28aff; }
    .stDownloadButton>button { color: #000; }
    .stTextArea textarea, .stDataFrame {
        background: rgba(162,138,255,0.15);
        color: #e0e0e0;
        border: 1px solid #a28aff;
        border-radius: 8px;
        padding: 10px;
        font-size: 0.9rem;
        font-family: Arial, sans-serif;
    }
    .stAlert { background: rgba(220,214,255,0.2); color: #e0e0e0; border-radius: 8px; font-size: 0.9rem; }
    .stText, table { color: #e0e0e0; font-family: Arial, sans-serif; font-size: 0.9rem; }
    .stDateInput label { color: #fff; font-size: 1.1rem; }
    .stDateInput>div>div>input { color: #000 !important; }

    div[data-baseweb="select"] {
        background: rgba(162,138,255,0.15) !important;
        border: 1px solid #a28aff !important;
        border-radius: 8px !important;
    }
    div[data-baseweb="select"] input {
        background: rgba(162,138,255,0.15) !important;
        color: #000 !important;
    }
    div[data-baseweb="select"] > div:first-child::before {
        content: none !important;
    }
    div[data-baseweb="select"] > div:first-child > div {
        padding-left: 12px !important;
    }

    @media (max-width: 600px) {
        h1 { font-size: 1.8rem; }
        .stFileUploader { margin-top: 10px; }
        .stFileUploader label { font-size: 0.9rem; }
        .stButton>button { padding: 8px 16px; font-size: 0.85rem; }
        .stTextArea textarea, .stDataFrame, .stAlert, .stText, table { font-size: 0.8rem; }
        .stDateInput label { font-size: 0.9rem; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# -------------------------------------------------
# Session state
# -------------------------------------------------
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame()
if "show_data" not in st.session_state:
    st.session_state.show_data = False
if "last_uploaded_file" not in st.session_state:
    st.session_state.last_uploaded_file = None

# -------------------------------------------------
# SUMMARY HELPERS
# -------------------------------------------------
def save_file_to_session(uploaded_file):
    filename = secure_filename(uploaded_file.name)
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(uploaded_file, dtype=str)
    elif ext in [".xls", ".xlsx"]:
        df = pd.read_excel(uploaded_file, dtype=str)
    else:
        raise ValueError("Only CSV / XLS / XLSX")
    df.columns = df.columns.str.strip()
    for v in ["CAMPAIGN", "campaign", "Campaign Name", "CAMPAIGN NAME"]:
        if v in df.columns:
            df = df.rename(columns={v: "Campaign"})
            break
    if not {"Collector Name", "Campaign"}.issubset(df.columns):
        raise ValueError("Need Collector Name & Campaign")
    if "Date" in df.columns:
        def norm(x):
            if pd.notnull(x) and str(x).strip():
                try:
                    return pd.to_datetime(x, errors="coerce").strftime("%m/%d/%Y")
                except:
                    pass
            return str(x).strip()
        df["Date"] = df["Date"].apply(norm)
    df["filename"] = filename
    df["upload_date"] = datetime.now().strftime("%Y-%m-%d")
    if st.session_state.df.empty:
        st.session_state.df = df
    else:
        st.session_state.df = pd.concat([st.session_state.df, df], ignore_index=True)
    return filename

def filter_records_by_date(start, end):
    s, e = start.strftime("%m/%d/%Y"), end.strftime("%m/%d/%Y")
    df = st.session_state.df.copy()
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()
    if "Date" not in df.columns:
        filtered = df
    else:
        filtered = df[df["Date"].between(s, e)] if start != end else df[df["Date"] == s]
    disp = filtered.copy()
    data_cols = [c for c in filtered.columns if c not in ["filename", "upload_date"]]
    disp = disp[["filename"] + data_cols + ["upload_date"]]
    return disp, filtered

def time_to_seconds(t):
    if pd.isna(t) or not str(t).strip():
        return 0.0
    try:
        parts = [float(p) if p.replace("-", "").replace(".", "").isdigit() else 0
                 for p in str(t).strip().split(":")]
        if len(parts) == 3:
            return max(parts[0], 0) * 3600 + max(parts[1], 0) * 60 + max(parts[2], 0)
        if len(parts) == 2:
            return max(parts[0], 0) * 60 + max(parts[1], 0)
    except:
        pass
    return 0.0

def seconds_to_time(s):
    s = int(float(s) if s else 0)
    h, s = divmod(s, 3600)
    m, s = divmod(s, 60)
    return f"{h}:{m:02d}:{s:02d}"

def calculate_agent_totals(df):
    metrics = ["Total Calls","Spent Time","Talk Time","Wait Time","Write Time","Pause Time"]
    avail = [c for c in metrics if c in df.columns]
    if not {"Collector Name"}.issubset(df.columns) or not avail:
        return pd.DataFrame()
    group = ["Collector Name","Campaign"] if "Campaign" in df.columns else ["Collector Name"]
    df = df.copy()
    df["Total Calls"] = pd.to_numeric(df.get("Total Calls",0),errors="coerce").fillna(0).astype(int)
    time_cols = [c for c in avail if c != "Total Calls"]
    for c in time_cols:
        df[f"{c}_sec"] = df[c].apply(time_to_seconds)
    sums = df.groupby(group)[["Total Calls"]+[f"{c}_sec" for c in time_cols]].sum().reset_index()
    for c in time_cols:
        sums[c] = sums[f"{c}_sec"].apply(seconds_to_time)
        sums.drop(columns=f"{c}_sec",inplace=True)
    if "Total Calls" in sums.columns and sums["Total Calls"].sum():
        for c in ["Talk Time","Wait Time","Write Time"]:
            if c in time_cols:
                sums[f"AVG {c}"] = sums.apply(
                    lambda r: seconds_to_time(time_to_seconds(r[c])/r["Total Calls"]) if r["Total Calls"] else "0:00:00",axis=1)
    cols = group + ["Total Calls"] + [c for c in ["Spent Time","Talk Time","Wait Time","Write Time","Pause Time"] if c in avail] \
           + [f"AVG {c}" for c in ["Talk Time","Wait Time","Write Time"] if f"AVG {c}" in sums.columns]
    return sums[cols]

def calculate_daily_agent_averages(df):
    metrics = ["Total Calls","Spent Time","Talk Time","Wait Time","Write Time","Pause Time"]
    avail = [c for c in metrics if c in df.columns]
    if not {"Collector Name","Date"}.issubset(df.columns) or not avail:
        return pd.DataFrame()
    group = ["Collector Name","Campaign"] if "Campaign" in df.columns else ["Collector Name"]
    daily = group + ["Date"]
    df = df.copy()
    df["Total Calls"] = pd.to_numeric(df.get("Total Calls",0),errors="coerce").fillna(0).astype(int)
    time_cols = [c for c in avail if c != "Total Calls"]
    for c in time_cols:
        df[f"{c}_sec"] = df[c].apply(time_to_seconds)
    day_avg = df.groupby(daily)[["Total Calls"]+[f"{c}_sec" for c in time_cols]].mean().reset_index()
    avg = day_avg.groupby(group)[["Total Calls"]+[f"{c}_sec" for c in time_cols]].mean().reset_index()
    avg["Total Calls"] = avg["Total Calls"].round().astype(int)
    for c in time_cols:
        avg[c] = avg[f"{c}_sec"].apply(seconds_to_time)
        avg.drop(columns=f"{c}_sec",inplace=True)
    if "Total Calls" in avg.columns and avg["Total Calls"].sum():
        for c in ["Talk Time","Wait Time","Write Time"]:
            if c in time_cols:
                avg[f"AVG {c}"] = avg.apply(
                    lambda r: seconds_to_time(time_to_seconds(r[c])/r["Total Calls"]) if r["Total Calls"] else "0:00:00",axis=1)
    cols = group + ["Total Calls"] + [c for c in ["Spent Time","Talk Time","Wait Time","Write Time","Pause Time"] if c in avail] \
           + [f"AVG {c}" for c in ["Talk Time","Wait Time","Write Time"] if f"AVG {c}" in avg.columns]
    return avg[cols]

def calculate_campaign_averages(df):
    metrics = ["Total Calls","Spent Time","Talk Time","Wait Time","Write Time","Pause Time"]
    avail = [c for c in metrics if c in df.columns]
    if "Campaign" not in df.columns or not avail:
        return pd.DataFrame()
    df = df.copy()
    df = df[df["Campaign"].notna() & (df["Campaign"].str.strip() != "")]
    if df.empty:
        return pd.DataFrame()
    df["Total Calls"] = pd.to_numeric(df.get("Total Calls",0),errors="coerce").fillna(0).astype(int)
    time_cols = [c for c in avail if c != "Total Calls"]
    for c in time_cols:
        df[f"{c}_sec"] = df[c].apply(time_to_seconds)
    avg = df.groupby("Campaign")[["Total Calls"]+[f"{c}_sec" for c in time_cols]].mean().reset_index()
    avg["Total Calls"] = avg["Total Calls"].round().astype(int)
    for c in time_cols:
        avg[c] = avg[f"{c}_sec"].apply(seconds_to_time)
        avg.drop(columns=f"{c}_sec",inplace=True)
    if "Total Calls" in avg.columns and avg["Total Calls"].sum():
        for c in ["Talk Time","Wait Time","Write Time"]:
            if c in time_cols:
                avg[f"AVG {c}"] = avg.apply(
                    lambda r: seconds_to_time(time_to_seconds(r[c])/r["Total Calls"]) if r["Total Calls"] else "0:00:00",axis=1)
    if {"Collector Name","Date"}.issubset(df.columns):
        agents = df.groupby(["Campaign","Date"])["Collector Name"].nunique().reset_index()
        avg_agents = agents.groupby("Campaign")["Collector Name"].mean().round().astype(int).reset_index()
        avg_agents.rename(columns={"Collector Name":"Average Number of Agents"},inplace=True)
        avg = avg.merge(avg_agents,on="Campaign",how="left").fillna({"Average Number of Agents":0})
    cols = ["Campaign","Total Calls"] + [c for c in ["Spent Time","Talk Time","Wait Time","Write Time","Pause Time"] if c in avail] \
           + [f"AVG {c}" for c in ["Talk Time","Wait Time","Write Time"] if f"AVG {c}" in avg.columns] \
           + (["Average Number of Agents"] if "Average Number of Agents" in avg.columns else [])
    return avg[cols]

def calculate_top_agents_lowest_spent_time_per_campaign(df):
    if not {"Campaign","Collector Name","Spent Time"}.issubset(df.columns):
        return pd.DataFrame(), pd.DataFrame()
    df = df.copy()
    df["Spent_sec"] = df["Spent Time"].apply(time_to_seconds)
    grp = df.groupby(["Campaign","Collector Name"])["Spent_sec"].mean().reset_index()
    grp = grp[grp["Spent_sec"]>0]
    if grp.empty:
        return pd.DataFrame(columns=["Agent","Average Spent Time"]), pd.DataFrame()
    top = grp.groupby("Campaign").apply(lambda x: x.nsmallest(5,"Spent_sec")).reset_index(drop=True)
    rows = []
    for camp in sorted(top["Campaign"].dropna().unique()):
        if not camp or not str(camp).strip():
            continue
        rows.append([camp,""])
        for _, r in top[top["Campaign"]==camp].head(5).iterrows():
            rows.append([r["Collector Name"],seconds_to_time(r["Spent_sec"])])
        rows.append(["",""])
    disp = pd.DataFrame(rows,columns=["Agent","Average Spent Time"])
    return disp, top

def calculate_bottom_10_agents_total_spent_time(df):
    if not {"Collector Name", "Spent Time"}.issubset(df.columns):
        return pd.DataFrame(columns=["Collector Name", "Total Spent Time"])
    df = df.copy()
    df["Spent_sec"] = df["Spent Time"].apply(time_to_seconds)
    df = df[df["Spent_sec"] > 0]
    if df.empty:
        return pd.DataFrame(columns=["Collector Name", "Total Spent Time"])
    totals = df.groupby("Collector Name")["Spent_sec"].sum().reset_index()
    bottom_10 = totals.nsmallest(10, "Spent_sec").copy()
    bottom_10["Total Spent Time"] = bottom_10["Spent_sec"].apply(seconds_to_time)
    bottom_10 = bottom_10[["Collector Name", "Total Spent Time"]]
    return bottom_10

# -------------------------------------------------
# Tabs
# -------------------------------------------------
tab1, tab2 = st.tabs(["PREDICTIVE COMPILER", "PREDICTIVE SUMMARY"])

# -------------------------------------------------
# TAB 1: PREDICTIVE COMPILER
# -------------------------------------------------
with tab1:
    st.title("PREDICTIVE COMPILER")
    uploaded_files = st.file_uploader(
        "Choose Excel files to merge",
        accept_multiple_files=True,
        type=["xlsx", "xls"],
        key="compiler_uploader"
    )

    if uploaded_files:
        UPLOAD_DIR = "tmp_compiler"
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        file_paths = []
        for f in uploaded_files:
            path = os.path.join(UPLOAD_DIR, f.name)
            with open(path, "wb") as out:
                out.write(f.getbuffer())
            file_paths.append(path)

        with st.spinner("Merging files..."):
            dfs = [pd.read_excel(p) for p in file_paths]
            df = pd.concat(dfs, ignore_index=True)

            for c in ["SNo.", "Pause Count"]:
                if c in df.columns:
                    df.drop(columns=c, inplace=True)

            if "Collector Name" not in df.columns:
                st.error("Missing 'Collector Name' column.")
            else:
                df = df[df["Collector Name"].notna() & (df["Collector Name"].str.strip() != "")]

                time_cols = [
                    "Spent Time", "Talk Time", "AVG Talk Time", "Wait Time",
                    "Average Wait Time", "Write Time", "AVG Write Time", "Pause Time"
                ]
                time_cols = [c for c in time_cols if c in df.columns]
                for c in time_cols:
                    df[c] = df[c].apply(
                        lambda x: sum(int(i) * 60**j for j, i in enumerate(reversed(str(x).split(":"))))
                        if pd.notna(x) and isinstance(x, str) else (int(x) if pd.notna(x) else 0)
                    )

                if "Total Calls" not in df.columns:
                    df["Total Calls"] = 1
                df["Total Calls"] = pd.to_numeric(df["Total Calls"], errors="coerce").fillna(1)

                agg = {c: "sum" for c in time_cols}
                agg["Total Calls"] = "sum"
                for c in df.columns:
                    if c not in time_cols + ["Collector Name", "Total Calls"]:
                        agg[c] = "first"
                grouped = df.groupby("Collector Name").agg(agg).reset_index()

                if "Total Calls" in grouped.columns and "Spent Time" in grouped.columns:
                    cols = grouped.columns.tolist()
                    cols.insert(cols.index("Spent Time"), cols.pop(cols.index("Total Calls")))
                    grouped = grouped[cols]

                avg_row = {"Collector Name": "Average"}
                for c in time_cols:
                    avg_row[c] = grouped[c].mean()
                avg_row["Total Calls"] = grouped["Total Calls"].mean()
                for c in grouped.columns:
                    if c not in avg_row:
                        avg_row[c] = None
                avg_df = pd.DataFrame([avg_row])[grouped.columns]
                result = pd.concat([grouped, avg_df], ignore_index=True)

                st.subheader("Merged Preview")
                preview = result.copy()
                for c in time_cols:
                    preview[c] = preview[c].apply(
                        lambda s: f"{int(s)//3600}:{int(s)%3600//60:02d}:{int(s)%60:02d}" if pd.notna(s) else "0:00:00"
                    )
                st.dataframe(preview, use_container_width=True)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    excel_df = result.copy()
                    for c in time_cols:
                        excel_df[c] = excel_df[c] / 86400
                    excel_df.to_excel(writer, index=False, sheet_name="Summary")
                    ws = writer.sheets["Summary"]
                    for c in time_cols:
                        col_idx = result.columns.get_loc(c) + 1
                        letter = get_column_letter(col_idx)
                        for r in range(2, len(result) + 2):
                            ws[f"{letter}{r}"].number_format = "[h]:mm:ss"
                            ws[f"{letter}{r}"].alignment = Alignment(horizontal="right")
                        ws[f"{letter}1"].alignment = Alignment(horizontal="right")
                    for i, col in enumerate(result.columns, 1):
                        if col != "Collector Name":
                            letter = get_column_letter(i)
                            for r in range(1, len(result) + 2):
                                ws[f"{letter}{r}"].alignment = Alignment(horizontal="right")
                output.seek(0)
                st.download_button(
                    "Download Merged Excel",
                    data=output,
                    file_name=f"Merged_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_compiler"
                )

        def cleanup():
            if os.path.exists(UPLOAD_DIR):
                shutil.rmtree(UPLOAD_DIR)
        import atexit
        atexit.register(cleanup)
    else:
        st.info("Upload Excel files above to merge and generate summary.")

# -------------------------------------------------
# TAB 2: PREDICTIVE SUMMARY
# -------------------------------------------------
with tab2:
    st.title("PREDICTIVE SUMMARY")

    st.header("Upload a File")
    uploaded_file = st.file_uploader(
        "Choose a file",
        type=["csv", "xls", "xlsx"],
        key="summary_uploader"
    )

    if uploaded_file is not None:
        current_file_name = uploaded_file.name
        if st.session_state.last_uploaded_file != current_file_name:
            try:
                filename = save_file_to_session(uploaded_file)
                st.session_state.last_uploaded_file = current_file_name
                st.success(f"Data from '{filename}' uploaded and stored successfully!")
            except Exception as e:
                st.error(f"Error: {str(e)}")

    st.header("Filter Data by Date Range")
    default_start_date = datetime.today().date()
    default_end_date = datetime.today().date()
    if "df" in st.session_state and not st.session_state.df.empty and "Date" in st.session_state.df.columns:
        try:
            valid_dates = pd.to_datetime(st.session_state.df["Date"], format="%m/%d/%Y", errors="coerce")
            valid_dates = valid_dates.dropna()
            if not valid_dates.empty:
                default_start_date = valid_dates.min().date()
                default_end_date = valid_dates.max().date()
        except:
            pass

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date", value=default_start_date)
    with col2:
        end_date = st.date_input("End Date", value=default_end_date)

    if st.button("Show All"):
        st.session_state.show_data = True

    if st.session_state.get("show_data", False):
        if start_date > end_date:
            st.error("Start date must be before or equal to end date.")
        else:
            _, raw_df = filter_records_by_date(start_date, end_date)
            if raw_df.empty:
                st.warning("No records found for the selected date range.")
            else:
                # TOTAL PER AGENT
                st.header("TOTAL PER AGENT")
                totals_df = calculate_agent_totals(raw_df)
                if not totals_df.empty:
                    all_agents = sorted(totals_df["Collector Name"].unique())
                    st.markdown("**Search for Agent (Total)**")
                    search_agent_total = st.multiselect(
                        "",
                        options=all_agents,
                        default=[],
                        key="search_total",
                        placeholder="Choose an option"
                    )
                    if search_agent_total:
                        filtered_totals = totals_df[totals_df["Collector Name"].isin(search_agent_total)]
                        st.dataframe(filtered_totals, use_container_width=True)
                    else:
                        st.info("Select one or more agents to view totals.")
                else:
                    st.warning("No data for totals.")

                # DAILY AVERAGE
                st.header("AVERAGE DAILY PER AGENT")
                daily_averages_df = calculate_daily_agent_averages(raw_df)
                if not daily_averages_df.empty:
                    all_agents_daily = sorted(daily_averages_df["Collector Name"].unique())
                    st.markdown("**Search for Agent (Daily Average)**")
                    search_agent_daily = st.multiselect(
                        "",
                        options=all_agents_daily,
                        default=[],
                        key="search_daily",
                        placeholder="Choose an option"
                    )
                    if search_agent_daily:
                        filtered_daily = daily_averages_df[daily_averages_df["Collector Name"].isin(search_agent_daily)]
                        st.dataframe(filtered_daily, use_container_width=True)
                    else:
                        st.info("Select one or more agents to view daily averages.")
                else:
                    st.warning("No data for daily averages.")

                # CAMPAIGN AVERAGE
                st.header("AVERAGE PER CAMPAIGN")
                campaign_averages_df = calculate_campaign_averages(raw_df)
                if not campaign_averages_df.empty:
                    st.dataframe(campaign_averages_df, use_container_width=True)
                else:
                    st.warning("No data for campaign averages.")

                # DOWNLOAD
                top_agents_df, top_agents_excel = calculate_top_agents_lowest_spent_time_per_campaign(raw_df)
                bottom_10_agents = calculate_bottom_10_agents_total_spent_time(raw_df)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_to_save = filtered_totals if 'filtered_totals' in locals() and not filtered_totals.empty else totals_df
                    if not df_to_save.empty:
                        df_to_save.to_excel(writer, index=False, sheet_name="Total per Agent")

                    df_daily = filtered_daily if 'filtered_daily' in locals() and not filtered_daily.empty else daily_averages_df
                    if not df_daily.empty:
                        df_daily.to_excel(writer, index=False, sheet_name="Average Daily per Agent")

                    if not campaign_averages_df.empty:
                        campaign_averages_df.to_excel(writer, index=False, sheet_name="Average per Campaign")

                    if not top_agents_excel.empty:
                        sheet_name = "Top Agents Lowest Avg Spent"
                        workbook = writer.book
                        worksheet = workbook.create_sheet(title=sheet_name)
                        row_idx = 1
                        for campaign in sorted(top_agents_excel["Campaign"].dropna().unique()):
                            if not campaign or not str(campaign).strip():
                                continue
                            worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                            cell = worksheet.cell(row=row_idx, column=1)
                            cell.value = campaign
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="left")
                            row_idx += 1
                            agents = top_agents_excel[top_agents_excel["Campaign"] == campaign].head(5)
                            for _, row in agents.iterrows():
                                worksheet.cell(row=row_idx, column=1).value = row["Collector Name"]
                                worksheet.cell(row=row_idx, column=2).value = seconds_to_time(row["Spent_sec"])
                                worksheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
                                worksheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left")
                                row_idx += 1
                            row_idx += 1
                        worksheet.column_dimensions["A"].width = 35
                        worksheet.column_dimensions["B"].width = 18

                    if not bottom_10_agents.empty:
                        bottom_10_agents.to_excel(writer, index=False, sheet_name="Bottom 10 Total Spent Time")

                output.seek(0)
                st.download_button(
                    label="Download Summary Excel",
                    data=output,
                    file_name=f"Summary_Data_{datetime.now():%m%d%y}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        if not st.session_state.df.empty:
            st.info("Click **'Show All'** to view summary data.")
        else:
            st.info("Upload a file to get started.")
